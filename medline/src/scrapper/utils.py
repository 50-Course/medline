import asyncio
import random
from contextlib import asynccontextmanager
from logging import Logger
from pathlib import Path
from typing import Annotated, Any, Callable, Dict, List, Optional

from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet
from playwright.async_api import ElementHandle, Locator, Page, async_playwright
from playwright.sync_api import BrowserContext

from .constants import USER_AGENTS as BROWSER_AGENTS
from .constants import url as BASE_URL


def get_random_user_agent() -> str:
    # gets a randomized browser agent
    return random.choice(BROWSER_AGENTS)


async def human_delay(min_=0.8, max_=2.5) -> None:
    await asyncio.sleep(random.uniform(min_, max_))


async def retry_with_backoff(coro: Callable, retries=3, delay=2):
    # extends the "core" concept of retry but with exponential backoff
    # and usable with any functtion
    for i in range(retries):
        try:
            return await coro()
        except Exception:
            await asyncio.sleep(delay * 2**i)
    raise RuntimeError("All retries failed")


async def goto_with_retry(
    page: Page, target_url: str, retries: int = 3, delay: int = 2
):
    # direct implement of page.goto but with a retry logic
    for i in range(retries):
        try:
            await page.goto(target_url, wait_until="domcontentloaded", timeout=10000)
            return
        except Exception as e:
            print(f"[WARN] Failed loading {BASE_URL}, retry {i + 1}/{retries}: {e}")
            await asyncio.sleep(delay * (2**i))


async def fallback_locator(
    page: Page,
    selectors: List[str],
    *,
    scope: Optional[Locator] = None,
    logger_func: Optional[Callable[[str], None]] = None,
    fallback_attrs: Optional[List[Dict[str, str]]] = None,
) -> Locator:
    """
    Try multiple selectors in order until one matches at least one element.
    Returns the first matching Locator.
    Raises ValueError if none match.
    """
    base = scope or page

    for selector in selectors:
        try:
            loc = base.locator(selector)
            if await loc.count() > 0:
                if logger_func:
                    logger_func(f"[→] Using selector: {selector}")
                return loc
        except Exception:
            continue

    # fallback to using attribute value for matching
    if fallback_attrs:
        for attr in fallback_attrs:
            tag_ = attr.get("tag", "*")
            key_ = attr.get("attr")
            value_ = attr.get("value")

            if key_ and value_:
                # build the expression
                expr = f"{tag_}[{key_}='{value_}']"

                try:
                    locator = base.locator(expr)
                    if await locator.count() > 0:
                        return locator
                except Exception:
                    # skip
                    continue

    raise ValueError(f"No selectors matched any elements: {selectors}")


@asynccontextmanager
async def browser_context(
    headless: bool = False,
    remote_debugging: bool = False,
    slow_mo: int = 50,
    user_agent: Optional[str] = None,
    viewport: Optional[dict] = None,
    bypass_csp: bool = False,
):
    async with async_playwright() as p:
        _args = [
            "--no-sandbox",
            "--disable-setuid-sandbox",
            "--disable-dev-shm-usage",
            "--disable-accelerated-2d-canvas",
            "--no-first-run",
            "--no-zygote",
            "--disable-gpu",
            "--disable-features=IsolateOrigins,site-per-process",
            "--window-size=1280,800",
            "--start-maximized",
        ]

        # browser = await p.chromium.launch(headless=headless, slow_mo=slow_mo, args=_args)
        browser = await p.firefox.launch(headless=headless, slow_mo=slow_mo)
        ctx = await browser.new_context(
            # user_agent=user_agent,
            # viewport=viewport or {"width": 1280, "height": 800},
            ignore_https_errors=True,
            java_script_enabled=True,
            bypass_csp=True,
        )
        try:
            yield ctx
        finally:
            await browser.close()


async def _get_rendered_html(page: Page, selector: str | None = None) -> BeautifulSoup:
    if selector:
        content = await page.query_selector(selector)
        if content:
            html = await content.inner_html()
            return BeautifulSoup(html, "html.parser")
        print(f"[ERROR] Selector {selector} did not return content.")
    html = await page.content()
    return BeautifulSoup(html, "html.parser")


def sanitize_sheet_name(name: str) -> str:
    """Ensure Excel sheet names are valid (max 31 chars, no special chars)."""
    invalid_chars = ["/", "\\", "*", "[", "]", ":", "?"]
    for char in invalid_chars:
        name = name.replace(char, "")
    return name[:31]


def auto_adjust_column_width(sheet: Worksheet):
    # auto-adjust column widths
    for col in sheet.columns:
        max_length = max(len(str(cell.value or "")) for cell in col)
        adjusted_width = max_length + 2
        col_letter = get_column_letter(col[0].column)
        sheet.column_dimensions[col_letter].width = adjusted_width


def write_subcategory_sheets(wb: Workbook, categories: List[Dict[str, Any]]):
    for category in categories:
        section = category["section"]
        for sub in category["subcategories"]:
            name = sanitize_sheet_name(sub["name"])
            index_entries = sub.get("index_entries", [])
            if not index_entries:
                continue

            ws = wb.create_sheet(title=name)
            headers = ["Section", "Title", "URL", "Image Src", "Image Alt"]
            ws.append(headers)

            for item in index_entries:
                ws.append(
                    [
                        section,
                        item.get("title", ""),
                        item.get("href", ""),
                        item.get("image_meta", {}).get("src", ""),
                        item.get("image_meta", {}).get("alt", ""),
                    ]
                )

            auto_adjust_column_width(ws)


def write_overview_sheet(wb: Workbook, categories: List[Dict[str, Any]]):
    overview = wb.active
    overview.title = "CATEGORIES CATALOG"
    headers = ["Category", "Subcategory", "URL"]
    overview.append(headers)

    for category in categories:
        category_name = category["section"]
        for sub in category["subcategories"]:
            overview.append([category_name, sub["name"], sub["url"]])

    auto_adjust_column_width(overview)


def write_category_to_excel(
    categories: List[Dict[str, Any]],
    filename: str = "scraped_expo_data.xlsx",
    output_dir: Path | None = None,
):
    if output_dir is None:
        base_dir = Path(__file__).resolve().parent
        output_dir = base_dir / "exports"
        output_dir.mkdir(parents=True, exist_ok=True)

    output_path = output_dir / filename

    wb = Workbook()
    write_overview_sheet(wb, categories)
    write_subcategory_sheets(wb, categories)
    write_products_to_excel(wb, categories)

    wb.save(output_path)
    print(f"[✓] Excel file saved to: {output_path}")


def write_products_to_excel(
    wb: Workbook,
    categories: list[dict[str, Any]],
):
    for category in categories:
        section = category["section"]
        for sub in category.get("subcategories", []):
            for entry in sub.get("index_entries", []):
                products = entry.get("products", [])

                if not products:
                    continue  # skip

                sheet_name = sanitize_sheet_name(entry["title"])
                ws = wb.create_sheet(title=sheet_name)

                headers = [
                    "Section",
                    "Subcategory",
                    "Entry Title",
                    "Product Name",
                    "Manufacturer",
                    "Price",
                    "Currency",
                    "Model",
                    "Features",
                    "Image Src",
                    "Link",
                ]
                ws.append(headers)

                for p in products:
                    ws.append(
                        [
                            section,
                            sub["name"],
                            entry["title"],
                            p.get("product_title"),
                            p.get("manufacturer_name"),
                            p.get("price"),
                            p.get("currency"),
                            p.get("product_model"),
                            ", ".join(p.get("features", [])),
                            p.get("tile_image_src"),
                            p.get("product_link"),
                        ]
                    )
                auto_adjust_column_width(ws)


def write_product_entry_to_excel(
    entry: Dict[str, Any], section: str, subcategory: str, output_dir: Path
):
    """Alternative function to write"""
    if not entry.get("products"):
        return

    title = sanitize_sheet_name(entry["title"])
    filename = f"{section}__{subcategory}__{title}.xlsx".replace(" ", "_")
    path = output_dir / filename

    wb = Workbook()
    ws = wb.active
    ws.title = "Products"

    headers = [
        "Product Name",
        "Manufacturer",
        "Price",
        "Currency",
        "Model",
        "Features",
        "Image Src",
        "Link",
    ]
    ws.append(headers)

    for p in entry["products"]:
        ws.append(
            [
                p.get("product_title"),
                p.get("manufacturer_name"),
                p.get("price"),
                p.get("currency"),
                p.get("product_model"),
                ", ".join(p.get("features", [])),
                p.get("tile_image_src"),
                p.get("product_link"),
            ]
        )

    auto_adjust_column_width(ws)
    wb.save(path)
    print(f"[✓] Saved: {path}")


async def extract_product_link_from_tile(tile: ElementHandle) -> str | None:
    """
    Extracts the actual product link from a product tile.
    Targets the <a> that wraps an <h3 class="short-name">.
    """
    import re

    try:
        # grabs the <h3> element that indicates a product link
        h3_el = await tile.query_selector("a[href] > h3.short-name")
        if h3_el:
            # walk back to the parent <a>
            parent_a = await h3_el.evaluate_handle("node => node.parentElement")
            if parent_a:
                href = await parent_a.get_attribute("href")
                if href and re.match(
                    r"^https://www\.medicalexpo\.com/prod/.+/product-\d+-\d+\.html$",
                    href,
                ):
                    return href
    except Exception as e:
        print(f"[ERROR] Could not extract product link: {e}")

    return None


async def safe_inner_text(locator: Locator) -> str | None:
    """Utility fn to help us safely access the inner text of a locator element without breaking the script"""
    try:
        if await locator.count() > 0:
            return (await locator.inner_text()).strip()
    except Exception:
        pass
    return None
